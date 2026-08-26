#!/usr/bin/env python3
"""Database housekeeping.

    python3 scripts/sync.py seed          # load the CSVs — ONCE, at setup
    python3 scripts/sync.py export        # database -> a dated .xlsx backup
    python3 scripts/sync.py pending       # community additions awaiting a check
    python3 scripts/sync.py verify 1043   # mark one verified
    python3 scripts/sync.py reject 1043   # remove one that isn't real (asks first)
    python3 scripts/sync.py add x.json    # write a researched entry (or - for stdin)

The database is the source of truth. `export` writes a dated snapshot for
safekeeping; nothing reads it back. Edit rows in the Supabase table editor,
not in a spreadsheet.

Needs SUPABASE_URL and SUPABASE_SERVICE_KEY in the environment or a .env file.
The service key bypasses Row Level Security, so it stays out of the browser.
"""
import os, sys, json, csv, re, pathlib, datetime, urllib.request, urllib.error, urllib.parse

ROOT = pathlib.Path(__file__).resolve().parent.parent
def load_env():
    f = ROOT/'.env'
    if f.exists():
        for line in f.read_text().splitlines():
            if '=' in line and not line.strip().startswith('#'):
                k,v=line.split('=',1); os.environ.setdefault(k.strip(), v.strip())
load_env()
URL = (os.environ.get('SUPABASE_URL') or '').rstrip('/')
KEY = os.environ.get('SUPABASE_SERVICE_KEY') or ''
if not URL or not KEY:
    sys.exit("Set SUPABASE_URL and SUPABASE_SERVICE_KEY (in the environment or .env).")

def req(method, path, body=None, extra=None):
    r = urllib.request.Request(URL+path, method=method,
        data=json.dumps(body).encode() if body is not None else None)
    r.add_header('apikey', KEY); r.add_header('Authorization','Bearer '+KEY)
    r.add_header('Content-Type','application/json')
    for k,v in (extra or {}).items(): r.add_header(k,v)
    try:
        with urllib.request.urlopen(r) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        sys.exit(f"{method} {path} -> {e.code}\n{e.read().decode()[:400]}")

def arr(v):
    v=(v or '').strip()
    return [x.strip('"') for x in v[1:-1].split(',') if x] if v.startswith('{') else []

def seed():
    for table, f in (('activities','seed_activities.csv'), ('events','seed_events.csv')):
        rows=[]
        for r in csv.DictReader(open(ROOT/'supabase'/f)):
            row={k:(v if v!='' else None) for k,v in r.items()}
            for k in ('tags','ages','season','conditions'):
                if k in row: row[k]=arr(row[k])
            for k in ('km','rating','lat','lng','id'):
                if row.get(k) is not None:
                    row[k]=float(row[k]) if k in ('km','lat','lng') else int(row[k])
            row['verified'] = str(row.get('verified')).lower()=='true'
            rows.append(row)
        for i in range(0,len(rows),100):
            req('POST', f'/rest/v1/{table}', rows[i:i+100],
                {'Prefer':'resolution=merge-duplicates,return=minimal'})
        print(f"pushed {len(rows)} -> {table}")
    # keep identity sequences ahead of the seeded ids
    print("NOTE: run this once in the SQL editor so new inserts don't collide:")
    print("  select setval(pg_get_serial_sequence('activities','id'), (select max(id) from activities));")
    print("  select setval(pg_get_serial_sequence('events','id'),     (select max(id) from events));")

def export():
    from openpyxl import load_workbook
    xl = ROOT/'JanJuc_WhatToDo_Database.xlsx'
    if not xl.exists(): sys.exit(f"Put the spreadsheet at {xl} first.")
    acts = req('GET','/rest/v1/activities?select=*&order=id')
    evs  = req('GET','/rest/v1/events?select=*&order=id')
    wb = load_workbook(xl)
    stamp = datetime.date.today().isoformat()
    out = ROOT/f'JanJuc_WhatToDo_Database.{stamp}.xlsx'
    ws = wb['Activities']
    hdr = {c.value: i+1 for i,c in enumerate(ws[1])}
    by_id = {a['id']: a for a in acts}
    updated = added = 0
    seen=set()
    for r in range(2, ws.max_row+1):
        i = ws.cell(r,1).value
        if i is None: continue
        a = by_id.get(int(i))
        if not a: continue
        seen.add(int(i))
        if ws.cell(r,11).value != a['description']:
            ws.cell(r,11).value = a['description']; updated += 1
    nxt = ws.max_row+1
    for a in acts:
        if a['id'] in seen: continue
        ws.cell(nxt,1).value=a['id']; ws.cell(nxt,2).value=a['name']; ws.cell(nxt,3).value=a['type']
        ws.cell(nxt,6).value=a.get('cost'); ws.cell(nxt,7).value=a.get('location')
        ws.cell(nxt,8).value=a.get('km');   ws.cell(nxt,11).value=a.get('description')
        ws.cell(nxt,12).value=a.get('url'); ws.cell(nxt,13).value=a.get('added_by')
        ws.cell(nxt,15).value=('UNVERIFIED — added on the site' if not a.get('verified') else None)
        ws.cell(nxt,16).value=', '.join(a.get('conditions') or ['any-weather'])
        nxt+=1; added+=1
    wb.save(out)
    print(f"pulled {len(acts)} activities, {len(evs)} events")
    print(f"  {added} new rows appended, {updated} descriptions refreshed")
    print(f"  written to {out.name} (the original is untouched)")

def pending():
    for t in ('activities','events'):
        rows = req('GET', f'/rest/v1/{t}?select=id,name,type,location,added_by,created_at&verified=eq.false&order=created_at.desc')
        print(f"\n{t}: {len(rows)} awaiting a check")
        for r in rows:
            print(f"  {r['id']:>6}  {r['name'][:44]:44} {str(r.get('added_by'))[:14]:14} {str(r.get('created_at'))[:10]}")

def verify(idn):
    for t in ('activities','events'):
        got = req('PATCH', f'/rest/v1/{t}?id=eq.{idn}', {'verified':True},
                  {'Prefer':'return=representation'})
        if got: print(f"verified {t} {idn}: {got[0]['name']}"); return
    print(f"no row with id {idn}")

def reject(idn, assume_yes=False):
    """Delete a community add. Refuses verified rows; confirms unless --yes."""
    for t in ('activities','events'):
        got = req('GET', f'/rest/v1/{t}?id=eq.{idn}&select=id,name,type,location,added_by,verified')
        if not got: continue
        r = got[0]
        if r['verified']:
            sys.exit(f"{t} {idn} '{r['name']}' is verified — un-verify it in Supabase first "
                     f"if you really mean to remove it.")
        print(f"{t} {idn}: {r['name']}")
        print(f"  type {r.get('type')}   location {r.get('location')}   added by {r.get('added_by')}")
        if not assume_yes and input("  delete permanently? [y/N] ").strip().lower() != 'y':
            print("  left alone."); return
        req('DELETE', f'/rest/v1/{t}?id=eq.{idn}')
        print(f"rejected {t} {idn}: {r['name']}")
        return
    print(f"no row with id {idn}")

# ── add ─────────────────────────────────────────────────────────────────────
# Write a researched entry straight in, bypassing the form. The service key
# ignores RLS, so this is the only path that may claim `verified` — which is
# why --verified insists on a source_note. Everything is checked here rather
# than left to the database, so a bad field in a batch of 200 names itself
# instead of failing an opaque insert halfway through.

ACTIVITY_COLS = {'name','types','tags','ages','cost','location','km','season','duration',
                 'description','url','rating','notes','conditions','lat','lng','daypart',
                 'added_by','verified','source_note'}
EVENT_COLS    = {'name','types','starts_on','ends_on','time_text','recurrence','venue',
                 'location','km','cost','ages','artist','genre','description','ticket_url',
                 'info_url','conditions','date_confidence','added_by','verified','source_note'}
EVENT_ONLY    = EVENT_COLS - ACTIVITY_COLS
URL_FIELDS    = ('url','ticket_url','info_url')

def vocab(table):
    return {r['name'] for r in req('GET', f'/rest/v1/{table}?select=name')}

def check(row, i, types, conds):
    """Return a list of complaints about one row. Empty list means it is fine."""
    bad = []
    where = f"row {i}" + (f" ({row['name']})" if row.get('name') else '')
    ev = bool(EVENT_ONLY & set(row)) or row.get('kind') == 'event'
    cols = EVENT_COLS if ev else ACTIVITY_COLS

    unknown = set(row) - cols - {'kind'}
    if unknown:
        # An event's link is info_url; writing `url` would be silently dropped.
        hint = ' (an event link is info_url or ticket_url)' if ev and 'url' in unknown else ''
        bad.append(f"{where}: no such field {sorted(unknown)}{hint}")
    if not str(row.get('name') or '').strip():
        bad.append(f"{where}: needs a name")
    # `types` is a list. A bare string is the old shape and is almost always a
    # hand-written row that predates the split, so name it rather than letting
    # Postgres turn 'gig' into the three-character array {g,i,g}.
    if 'types' in row:
        if isinstance(row['types'], str):
            bad.append(f"{where}: types must be a list — write [\"{row['types']}\"], not \"{row['types']}\"")
        else:
            for t in (row['types'] or []):
                if t not in types:
                    bad.append(f"{where}: type '{t}' is not one of the {len(types)} allowed")
    for c in (row.get('conditions') or []):
        if c not in conds: bad.append(f"{where}: condition '{c}' is not in the vocabulary")
    if not ev and row.get('cost') not in (None,'Free','Cheap','Moderate','Splurge'):
        bad.append(f"{where}: cost '{row['cost']}' must be Free/Cheap/Moderate/Splurge")
    if not ev and row.get('daypart') not in (None,'day','night','both'):
        bad.append(f"{where}: daypart must be day/night/both")
    if ev and row.get('recurrence') not in (None,'none','weekly','fortnightly','monthly','annual'):
        bad.append(f"{where}: recurrence '{row['recurrence']}' is not allowed")
    if ev and row.get('date_confidence') not in (None,'high','medium','low'):
        bad.append(f"{where}: date_confidence must be high/medium/low")
    if ev and row.get('starts_on') and not re.match(r'^\d{4}-\d{2}-\d{2}$', str(row['starts_on'])):
        bad.append(f"{where}: starts_on must be YYYY-MM-DD")
    for f in URL_FIELDS:
        u = row.get(f)
        if u and not str(u).startswith(('http://','https://')):
            bad.append(f"{where}: {f} must start with http:// or https://")
        if u and 'maps.app.goo.gl' in str(u):
            bad.append(f"{where}: {f} is a maps.app.goo.gl link — those get fabricated, use a real one")
        # A Google Maps *search* url is not a link to the place, it is a link to
        # a search box. It reads as a filled-in field, so nobody goes back for
        # the real one, and it is a missing pin forever. Two research passes have
        # now written these while the rule sat in prose in RESEARCH_RULES.md —
        # round one called it "a Maps-search placeholder used per policy". There
        # is no such policy. Null is the honest value for a venue with no site.
        if u and re.search(r'google\.[a-z.]+/maps/search/', str(u)):
            bad.append(f"{where}: {f} is a Google Maps search link, not a link to the place. "
                       f"There is no placeholder policy — leave the url null if the venue "
                       f"has no site of its own.")
    if row.get('verified') and not str(row.get('source_note') or '').strip():
        bad.append(f"{where}: claiming verified without a source_note — say where it came from")
    # 0.01 degree is 1.1km, which on this coast is often open water. /admin has
    # refused these since it was built; this path did not, so a 3-decimal pin
    # walked straight in on 26 Aug 2026. Two write paths, one rule.
    for f in ('lat','lng'):
        v = row.get(f)
        if v is None: continue
        dp = len(str(v).split('.')[1]) if '.' in str(v) else 0
        if dp < 4:
            bad.append(f"{where}: {f} {v} has {dp} decimal places — under four is a "
                       f"guess, not a coordinate. Geocode it or leave both null. "
                       f"(A real match can land on a round number: OSM has the 18th "
                       f"Amendment Bar at exactly -38.1480000. If that is what this is, "
                       f"say so in source_note and write it through /admin.)")
    return bad

def add(path, verified=False, dry=False, force=False):
    raw = sys.stdin.read() if path == '-' else pathlib.Path(path).read_text()
    try: doc = json.loads(raw)
    except json.JSONDecodeError as e: sys.exit(f"that is not valid JSON: {e}")
    rows = doc if isinstance(doc, list) else [doc]

    types, conds = vocab('types'), vocab('conditions')
    for r in rows:
        r.setdefault('added_by','Research')
        if verified: r['verified'] = True

    bad = [c for i,r in enumerate(rows,1) for c in check(r, i, types, conds)]
    if bad:
        print("nothing written — fix these first:"); [print("  •",b) for b in bad]; sys.exit(1)

    # The same festival landing twice under two spellings is how this database
    # got into trouble before. Say so; --force if it really is a separate thing.
    clashes = []
    for r in rows:
        q = urllib.parse.quote(r['name'].strip())
        for t in ('activities','events'):
            for hit in req('GET', f'/rest/v1/{t}?select=id,name,verified&name=ilike.{q}'):
                clashes.append(f"  • {r['name']} looks like {t} {hit['id']} "
                               f"'{hit['name']}' (verified={hit['verified']})")
    if clashes and not force:
        print("already in the database:"); [print(c) for c in clashes]
        sys.exit("nothing written. --force if these really are different things.")

    for r in rows:
        ev = bool(EVENT_ONLY & set(r)) or r.pop('kind', None) == 'event'
        r.pop('kind', None)
        t = 'events' if ev else 'activities'
        if dry:
            print(f"would add to {t}: {r['name']}"); continue
        got = req('POST', f'/rest/v1/{t}', r, {'Prefer':'return=representation'})
        row = got[0] if got else {}
        print(f"added {t} {row.get('id')}: {row.get('name')}"
              f"  verified={row.get('verified')}")
    if dry: print(f"dry run — {len(rows)} row(s) checked, nothing written")


cmd = sys.argv[1] if len(sys.argv)>1 else ''
if   cmd=='seed': seed()
elif cmd=='export': export()
elif cmd=='pending': pending()
elif cmd=='verify' and len(sys.argv)>2: verify(sys.argv[2])
elif cmd=='reject' and len(sys.argv)>2: reject(sys.argv[2], '--yes' in sys.argv)
elif cmd=='add' and len(sys.argv)>2: add(sys.argv[2], '--verified' in sys.argv,
                                         '--dry-run' in sys.argv, '--force' in sys.argv)
else: print(__doc__)
